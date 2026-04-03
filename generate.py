from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
import tempfile
import os
import shutil
import zipfile
import json
from pathlib import Path
from decimal import Decimal
import subprocess
import openpyxl
from io import BytesIO

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

ALLOWED_EXTENSIONS = {'docx', 'xlsx'}

def allowed_file(filename, ext):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_float_increment(start_str):
    """Parse start number (e.g., '2604.01') and return increment function"""
    try:
        start = Decimal(start_str)
        current = [start]
        
        def get_next():
            result = current[0]
            current[0] += Decimal('0.01')
            return str(result)
        
        return get_next
    except:
        try:
            start = int(start_str.split('.')[0])
            current = [start]
            
            def get_next():
                result = current[0]
                current[0] += 1
                return str(result)
            
            return get_next
        except:
            return lambda: start_str

def extract_docx(docx_path, extract_dir):
    """Extract DOCX (ZIP) to directory"""
    with zipfile.ZipFile(docx_path, 'r') as zip_ref:
        zip_ref.extractall(extract_dir)

def find_and_replace_in_xml(file_path, replacements):
    """Find and replace text in XML file"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    for old, new in replacements.items():
        content = content.replace(old, new)
    
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write(content)

def repack_docx(extract_dir, output_docx):
    """Repack DOCX from extracted directory"""
    with zipfile.ZipFile(output_docx, 'w', zipfile.ZIP_DEFLATED) as docx:
        for root, dirs, files in os.walk(extract_dir):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, extract_dir)
                docx.write(file_path, arcname)

def convert_docx_to_pdf(docx_path, pdf_path):
    """Convert DOCX to PDF using LibreOffice"""
    try:
        output_dir = os.path.dirname(pdf_path)
        
        result = subprocess.run([
            'libreoffice', '--headless', '--convert-to', 'pdf',
            '--outdir', output_dir,
            docx_path
        ], capture_output=True, timeout=30)
        
        temp_pdf = os.path.join(output_dir, Path(docx_path).stem + '.pdf')
        if os.path.exists(temp_pdf):
            if temp_pdf != pdf_path:
                shutil.move(temp_pdf, pdf_path)
            return True
        
        return False
    except:
        return False

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

@app.route('/api/generate', methods=['POST'])
def generate():
    """Generate certificates from uploaded files"""
    
    try:
        # Check files
        if 'docx' not in request.files or 'excel' not in request.files:
            return jsonify({'error': 'Missing files'}), 400
        
        docx_file = request.files['docx']
        excel_file = request.files['excel']
        
        if not docx_file or not excel_file:
            return jsonify({'error': 'Empty files'}), 400
        
        if not allowed_file(docx_file.filename, 'docx') or not allowed_file(excel_file.filename, 'xlsx'):
            return jsonify({'error': 'Invalid file types'}), 400
        
        # Get form data
        course_name = request.form.get('courseName', '').strip()
        issue_date = request.form.get('issueDate', '').strip()
        study_load = request.form.get('studyLoad', '').strip()
        start_number = request.form.get('startNumber', '').strip()
        
        if not all([course_name, issue_date, study_load, start_number]):
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Create temp directory
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save uploaded files
            docx_path = os.path.join(temp_dir, 'template.docx')
            excel_path = os.path.join(temp_dir, 'participants.xlsx')
            output_dir = os.path.join(temp_dir, 'output')
            
            docx_file.save(docx_path)
            excel_file.save(excel_path)
            os.makedirs(output_dir, exist_ok=True)
            
            # Read Excel
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.lower()] = col_idx
            
            voornaam_col = headers.get('voornaam')
            achternaam_col = headers.get('achternaam')
            
            if not voornaam_col or not achternaam_col:
                return jsonify({'error': 'Excel must contain Voornaam and Achternaam columns'}), 400
            
            # Collect participants
            participants = []
            for row_idx in range(2, ws.max_row + 1):
                voornaam = ws.cell(row_idx, voornaam_col).value
                achternaam = ws.cell(row_idx, achternaam_col).value
                
                if voornaam and achternaam:
                    participants.append(f"{voornaam} {achternaam}")
            
            if not participants:
                return jsonify({'error': 'No participants found'}), 400
            
            # Generate certificates
            number_generator = get_float_increment(start_number)
            pdfs = []
            
            for idx, participant in enumerate(participants, 1):
                try:
                    with tempfile.TemporaryDirectory() as cert_temp:
                        # Extract template
                        template_dir = os.path.join(cert_temp, 'template')
                        extract_docx(docx_path, template_dir)
                        
                        # Replace in document.xml
                        doc_xml = os.path.join(template_dir, 'word', 'document.xml')
                        
                        replacements = {
                            'Gloria Chang Drop': participant,
                            'an introduction to Nordoff-Robbins Music Therapy': course_name,
                            '5 maart 2026': issue_date,
                            '16 uur': f"{study_load} uur",
                            '2603.01': number_generator()
                        }
                        
                        find_and_replace_in_xml(doc_xml, replacements)
                        
                        # Repack DOCX
                        modified_docx = os.path.join(cert_temp, f'cert_{idx}.docx')
                        repack_docx(template_dir, modified_docx)
                        
                        # Convert to PDF
                        safe_name = participant.replace(' ', '_').replace('/', '_')
                        pdf_path = os.path.join(output_dir, f'{safe_name}_{idx:03d}.pdf')
                        
                        if convert_docx_to_pdf(modified_docx, pdf_path):
                            pdfs.append(pdf_path)
                
                except Exception as e:
                    print(f"Error processing {participant}: {e}")
                    continue
            
            if not pdfs:
                return jsonify({'error': 'Failed to generate any certificates'}), 500
            
            # Create ZIP with all PDFs
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for pdf in pdfs:
                    arcname = os.path.basename(pdf)
                    zip_file.write(pdf, arcname)
            
            zip_buffer.seek(0)
            
            return send_file(
                zip_buffer,
                mimetype='application/zip',
                as_attachment=True,
                download_name='certificaten.zip'
            )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=False)
